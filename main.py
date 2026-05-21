#!/usr/bin/env python3
"""CSV to XLSX Converter + Facebook Audience Formatter.

Two functions, two tabs:

  1. Converter — RFC 4180 CSV (500k+ rows, 670MB+) → XLSX, treating every
     cell as text so UUIDs, phone numbers, leading-zero zips, and ranges
     like "$20,000 to $44,999" are preserved exactly.

  2. Facebook Audience Formatter — turn a raw or converted contact list
     into a Meta Custom Audience-ready CSV/XLSX. Splits multi-value
     email/phone columns into EMAIL/EMAIL_2/EMAIL_3 + PHONE/PHONE_2,
     emits SHA256-hashed siblings alongside plaintext (both toggleable),
     normalizes per Meta's hashing spec.

Architecture:
  - PyQt6 UI on the main thread (drag/drop, preview, progress).
  - Polars on a worker thread for parsing the converter input; pandas
    chunked fallback if Polars fails or isn't available on this platform.
  - xlsxwriter in constant_memory mode for the write side.
  - Formatter uses stdlib csv + openpyxl read-only iteration so it
    streams instead of slurping; SHA256 hashing happens row-by-row.
"""

import csv
import hashlib
import os
import re
import subprocess
import sys
import traceback
from pathlib import Path
from typing import Iterator, Optional

from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QDragEnterEvent, QDropEvent
from PyQt6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFormLayout,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpinBox,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

EXCEL_MAX_ROWS = 1_048_576  # Excel's hard per-sheet limit.

# Polars is the fast path; pandas is the safety net. We probe both at
# import time so the worker thread can decide per-conversion without
# paying for it again.
try:
    import polars as pl

    POLARS_AVAILABLE = True
except Exception:
    POLARS_AVAILABLE = False

import pandas as pd
import xlsxwriter

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


def human_size(n: float) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} TB"


# =====================================================================
# Converter: CSV -> XLSX
# =====================================================================


class RowCountWorker(QThread):
    """Count rows and pull a 5-row preview without slurping the file."""

    done = pyqtSignal(int, list, list)  # count, columns, preview rows
    failed = pyqtSignal(str)

    def __init__(self, path: str):
        super().__init__()
        self.path = path

    def run(self) -> None:
        try:
            if POLARS_AVAILABLE:
                lf = pl.scan_csv(self.path, infer_schema_length=0)
                # pl.len() over a lazy frame counts rows without materializing data.
                count = lf.select(pl.len()).collect().item()
                preview = lf.head(5).collect()
                cols = list(preview.columns)
                rows = [list(r) for r in preview.iter_rows()]
                self.done.emit(int(count), cols, rows)
                return

            # Pandas fallback: preview with read_csv(nrows=5), then a cheap
            # newline count. Note this isn't exact for CSVs with quoted
            # newlines, but the progress bar uses it only as an estimate.
            preview_df = pd.read_csv(
                self.path,
                dtype=str,
                nrows=5,
                na_filter=False,
                keep_default_na=False,
            )
            cols = list(preview_df.columns)
            rows = preview_df.values.tolist()
            count = 0
            with open(self.path, "rb") as f:
                for _ in f:
                    count += 1
            self.done.emit(max(0, count - 1), cols, rows)
        except Exception as e:
            self.failed.emit(str(e))


class ConvertWorker(QThread):
    progress = pyqtSignal(int, int)  # processed, total
    status = pyqtSignal(str)
    warning = pyqtSignal(str)
    failed = pyqtSignal(str)
    done = pyqtSignal(str)  # output path

    def __init__(self, input_path: str, output_path: str, total_rows: int):
        super().__init__()
        self.input_path = input_path
        self.output_path = output_path
        self.total_rows = max(total_rows, 1)

    def run(self) -> None:
        try:
            # Try Polars first because:
            #   - infer_schema_length=0 forces every column to stay Utf8,
            #     so UUIDs, phone numbers, zips with leading zeros, and
            #     ranges like "$20,000 to $44,999" never get auto-typed.
            #   - it handles RFC 4180 quoted fields with embedded commas,
            #     quotes, and newlines correctly out of the box.
            # If Polars fails for any reason (rare schema bug, encoding
            # quirk, missing wheel on this platform), drop into the pandas
            # chunked path. It's slower but battle-tested.
            if POLARS_AVAILABLE:
                try:
                    self._convert_with_polars()
                    return
                except MemoryError:
                    raise
                except Exception as e:
                    self.warning.emit(
                        f"Polars path failed ({e}). Falling back to pandas."
                    )
            self._convert_with_pandas()
        except MemoryError:
            self.failed.emit(
                "Out of memory while converting. Close other apps and retry, "
                "or split the CSV into smaller files first."
            )
        except Exception as e:
            tb = traceback.format_exc(limit=3)
            self.failed.emit(f"{e}\n\n{tb}")

    def _convert_with_polars(self) -> None:
        self.status.emit("Reading CSV with Polars (all columns as text)...")
        df = pl.read_csv(
            self.input_path,
            infer_schema_length=0,
            ignore_errors=False,
            truncate_ragged_lines=False,
            try_parse_dates=False,
            null_values=[],  # don't treat any literal as null; keep empty as empty
        )
        self.status.emit(
            f"Loaded {df.height:,} rows x {df.width} columns. Writing XLSX..."
        )
        # Update total now that we know it exactly.
        self.total_rows = max(df.height, 1)
        self.progress.emit(0, self.total_rows)
        self._write_xlsx(df.iter_rows(), list(df.columns), df.height)

    def _convert_with_pandas(self) -> None:
        self.status.emit("Reading CSV with pandas (chunked)...")
        chunk_iter = pd.read_csv(
            self.input_path,
            dtype=str,
            na_filter=False,
            keep_default_na=False,
            chunksize=50_000,
            engine="c",
            on_bad_lines="warn",
        )
        first = next(chunk_iter)
        cols = list(first.columns)

        def row_gen():
            for row in first.itertuples(index=False, name=None):
                yield row
            for chunk in chunk_iter:
                for row in chunk.itertuples(index=False, name=None):
                    yield row

        self._write_xlsx(row_gen(), cols, self.total_rows)

    def _write_xlsx(self, row_iter, columns: list, total_rows: int) -> None:
        # constant_memory: only the current row lives in RAM, prior rows
        # are flushed to disk. Required for 500k+ row workbooks.
        workbook = xlsxwriter.Workbook(
            self.output_path,
            {
                "constant_memory": True,
                "strings_to_formulas": False,
                "strings_to_urls": False,
                "strings_to_numbers": False,
            },
        )
        header_fmt = workbook.add_format({"bold": True})

        sheet_idx = 1
        sheet = workbook.add_worksheet(f"Data_{sheet_idx}")
        self._write_header(sheet, columns, header_fmt)

        max_data_rows = EXCEL_MAX_ROWS - 1  # leave row 0 for the header
        row_in_sheet = 1
        processed = 0
        emit_every = max(1, min(5000, total_rows // 100 or 1))

        for row in row_iter:
            if row_in_sheet > max_data_rows:
                sheet_idx += 1
                sheet = workbook.add_worksheet(f"Data_{sheet_idx}")
                self._write_header(sheet, columns, header_fmt)
                row_in_sheet = 1
                self.warning.emit(
                    f"Per-sheet row limit reached, continuing in sheet Data_{sheet_idx}."
                )

            for c, val in enumerate(row):
                # None / "" → leave the cell empty rather than writing
                # the strings "NaN" or "None".
                if val is None or val == "":
                    continue
                sheet.write_string(row_in_sheet, c, str(val))

            row_in_sheet += 1
            processed += 1

            if processed % emit_every == 0:
                self.progress.emit(processed, max(total_rows, processed))
                self.status.emit(
                    f"Wrote {processed:,} / {max(total_rows, processed):,} rows..."
                )

        workbook.close()
        self.progress.emit(processed, max(total_rows, processed))
        self.status.emit(f"Done. Wrote {processed:,} rows to {self.output_path}")
        self.done.emit(self.output_path)

    @staticmethod
    def _write_header(sheet, columns: list, fmt) -> None:
        for c, name in enumerate(columns):
            sheet.write_string(0, c, "" if name is None else str(name), fmt)


# =====================================================================
# Facebook Audience Formatter
# =====================================================================

# Facebook Custom Audience field names per Meta's hashing spec.
# Reference: https://www.facebook.com/business/help/606443329504150
#
# EMAIL and PHONE are handled as multi-source (a row can pull from any
# number of input columns); everything else is a single-source mapping.
FB_SINGLE_FIELDS = [
    "EXTERN_ID",
    "FN",
    "LN",
    "ZIP",
    "CT",
    "ST",
    "COUNTRY",
    "DOB",
    "GEN",
]

# Heuristics for auto-mapping common source column names to FB fields.
# Matched case-insensitively; exact match first, substring fallback.
AUTO_MAP = {
    "EXTERN_ID": ["uuid", "id", "extern_id", "external_id", "person_id"],
    "FN": ["first_name", "fn", "firstname", "given_name"],
    "LN": ["last_name", "ln", "lastname", "surname", "family_name"],
    "ZIP": ["zip", "zip_code", "postal_code", "postcode"],
    "CT": ["city", "ct"],
    "ST": ["state", "st", "region"],
    "COUNTRY": ["country", "country_code", "iso_country"],
    "DOB": ["dob", "date_of_birth", "birthdate", "birth_date"],
    "GEN": ["gender", "sex", "gen"],
}

# Substring patterns for auto-checking multi-source email / phone lists.
EMAIL_COLUMN_PATTERNS = ["email", "e-mail", "e_mail"]
PHONE_COLUMN_PATTERNS = ["phone", "mobile", "cell", "msisdn"]


def auto_map_columns(fb_field: str, columns: list) -> str:
    """Pick the most likely source column for fb_field, or '' if none fit."""
    patterns = AUTO_MAP.get(fb_field, [])
    lower = [c.lower() for c in columns]
    for pat in patterns:
        for i, c in enumerate(lower):
            if c == pat:
                return columns[i]
    for pat in patterns:
        for i, c in enumerate(lower):
            if pat in c:
                return columns[i]
    return ""


# --- normalization helpers (Meta's hashing rules) ---------------------


def norm_email(s: str) -> str:
    return s.strip().lower() if s else ""


def norm_phone(s: str, country_prefix: str = "1") -> str:
    """Strip everything but digits. If 10 digits (US local), prepend the
    given country prefix so the value is in E.164 form minus the '+'."""
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if len(digits) == 10 and country_prefix:
        digits = country_prefix + digits
    return digits


def norm_name(s: str) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", "", s.strip().lower())


def norm_zip(s: str) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    # US ZIP+4: take just the 5-digit prefix.
    if "-" in s:
        s = s.split("-")[0]
    return s


def norm_city(s: str) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", "", s.strip().lower())


def norm_state(s: str) -> str:
    if not s:
        return ""
    return s.strip().lower()[:2]


def norm_country(s: str) -> str:
    if not s:
        return ""
    return s.strip().lower()[:2]


def norm_gender(s: str) -> str:
    if not s:
        return ""
    c = s.strip().lower()[:1]
    return c if c in ("m", "f") else ""


def parse_dob(s: str) -> tuple:
    """Return (Y, M, D) strings from a DOB; ('','','') if unparseable.
    Accepts YYYY-MM-DD, MM/DD/YYYY, M/D/YY, YYYYMMDD, and dot-separated."""
    if not s:
        return "", "", ""
    s = s.strip()
    m = re.match(r"(\d{4})[-/\.](\d{1,2})[-/\.](\d{1,2})", s)
    if m:
        return m.group(1), m.group(2).zfill(2), m.group(3).zfill(2)
    m = re.match(r"(\d{1,2})[-/\.](\d{1,2})[-/\.](\d{2,4})", s)
    if m:
        y = m.group(3)
        if len(y) == 2:
            y = "19" + y if int(y) > 30 else "20" + y
        return y, m.group(1).zfill(2), m.group(2).zfill(2)
    m = re.match(r"(\d{4})(\d{2})(\d{2})", s)
    if m:
        return m.group(1), m.group(2), m.group(3)
    return "", "", ""


def sha256_hex(s: str) -> str:
    return hashlib.sha256(s.encode("utf-8")).hexdigest() if s else ""


# --- input file reading (CSV + XLSX, streaming) ------------------------


def read_input_columns(path: str) -> list:
    """Return the header columns of a CSV or XLSX without reading the body."""
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            return next(reader, [])
    if ext in (".xlsx", ".xlsm"):
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required to read XLSX files.")
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            # Use the first Data_* sheet if present (matches the converter's
            # multi-sheet output), otherwise the active sheet.
            sheets = [s for s in wb.sheetnames if s.startswith("Data_")] or [
                wb.active.title
            ]
            ws = wb[sheets[0]]
            header_row = next(ws.iter_rows(values_only=True), ())
            return [str(c) if c is not None else "" for c in header_row]
        finally:
            wb.close()
    raise ValueError(f"Unsupported file type: {ext}")


def iter_input_rows(path: str) -> Iterator[dict]:
    """Stream rows from CSV or XLSX as dicts of {column: string}.
    For XLSX, all sheets named Data_* are read in order; if none exist,
    only the active sheet is read."""
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield {k: ("" if v is None else str(v)) for k, v in row.items()}
        return
    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = [s for s in wb.sheetnames if s.startswith("Data_")] or [
                wb.active.title
            ]
            for sname in sheets:
                ws = wb[sname]
                header = None
                for row in ws.iter_rows(values_only=True):
                    if header is None:
                        header = [str(c) if c is not None else "" for c in row]
                        continue
                    d = {}
                    for i, c in enumerate(row):
                        if i < len(header):
                            d[header[i]] = "" if c is None else str(c)
                    yield d
        finally:
            wb.close()
        return
    raise ValueError(f"Unsupported file type: {ext}")


def count_input_rows(path: str) -> int:
    """Approximate row count for progress reporting. Exact for XLSX,
    approximate for CSVs that contain quoted newlines."""
    ext = Path(path).suffix.lower()
    if ext == ".csv":
        count = 0
        with open(path, "rb") as f:
            for _ in f:
                count += 1
        return max(0, count - 1)
    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheets = [s for s in wb.sheetnames if s.startswith("Data_")] or [
                wb.active.title
            ]
            total = 0
            for sname in sheets:
                ws = wb[sname]
                total += max(0, (ws.max_row or 0) - 1)
            return total
        finally:
            wb.close()
    return 0


# --- formatter worker --------------------------------------------------


class AudienceFormatWorker(QThread):
    progress = pyqtSignal(int, int)  # processed, total
    status = pyqtSignal(str)
    warning = pyqtSignal(str)
    failed = pyqtSignal(str)
    done = pyqtSignal(str)

    def __init__(
        self,
        input_path: str,
        output_path: str,
        mapping: dict,
        options: dict,
        total_rows: int,
    ):
        super().__init__()
        self.input_path = input_path
        self.output_path = output_path
        self.mapping = mapping
        self.options = options
        self.total_rows = max(total_rows, 1)

    def run(self) -> None:
        try:
            self._run()
        except Exception as e:
            tb = traceback.format_exc(limit=3)
            self.failed.emit(f"{e}\n\n{tb}")

    def _run(self) -> None:
        cols = self._build_output_columns()
        out_fmt = self.options["output_format"]
        self.status.emit(f"Writing {out_fmt.upper()} → {self.output_path}")
        if out_fmt == "csv":
            self._write_csv(cols)
        else:
            self._write_xlsx(cols)
        self.done.emit(self.output_path)

    def _build_output_columns(self) -> list:
        """Build the output schema. A column is only emitted if its source
        is actually mapped — that's how the user controls how wide the
        output is. EMAIL/PHONE expand to N columns based on the max
        spinners; the others are single columns (DOB expands to 3)."""
        opts = self.options
        m = self.mapping
        max_e = opts["max_emails"]
        max_p = opts["max_phones"]
        plain = opts["include_plaintext"]
        sha = opts["include_sha256"]

        cols: list = []
        if m.get("EXTERN_ID"):
            cols.append("EXTERN_ID")
        if m.get("FN"):
            cols.append("FN")
        if m.get("LN"):
            cols.append("LN")

        if m.get("EMAIL_SOURCES"):
            for i in range(max_e):
                suffix = "" if i == 0 else f"_{i+1}"
                if plain:
                    cols.append(f"EMAIL{suffix}")
                if sha:
                    cols.append(f"EMAIL_SHA256{suffix}")
        if m.get("PHONE_SOURCES"):
            for i in range(max_p):
                suffix = "" if i == 0 else f"_{i+1}"
                if plain:
                    cols.append(f"PHONE{suffix}")
                if sha:
                    cols.append(f"PHONE_SHA256{suffix}")

        if m.get("ZIP"):
            cols.append("ZIP")
        if m.get("CT"):
            cols.append("CT")
        if m.get("ST"):
            cols.append("ST")
        if m.get("COUNTRY"):
            cols.append("COUNTRY")
        if m.get("DOB"):
            cols += ["DOBY", "DOBM", "DOBD"]
        if m.get("GEN"):
            cols.append("GEN")
        return cols

    def _build_row(self, src: dict, cols: list) -> dict:
        m = self.mapping
        opts = self.options
        max_e = opts["max_emails"]
        max_p = opts["max_phones"]
        plain = opts["include_plaintext"]
        sha = opts["include_sha256"]
        sep = opts["separator"]
        prefix = opts["country_prefix"]

        out = {c: "" for c in cols}

        if m.get("EXTERN_ID"):
            out["EXTERN_ID"] = (src.get(m["EXTERN_ID"]) or "").strip()
        if m.get("FN"):
            out["FN"] = norm_name(src.get(m["FN"], ""))
        if m.get("LN"):
            out["LN"] = norm_name(src.get(m["LN"], ""))

        # Email: pool across every selected source column, dedupe, cap.
        if m.get("EMAIL_SOURCES"):
            collected: list = []
            for col in m["EMAIL_SOURCES"]:
                raw = src.get(col, "") or ""
                for p in raw.split(sep):
                    n = norm_email(p)
                    if n:
                        collected.append(n)
            seen = set()
            uniq = []
            for p in collected:
                if p not in seen:
                    seen.add(p)
                    uniq.append(p)
            uniq = uniq[:max_e]
            for i, val in enumerate(uniq):
                suffix = "" if i == 0 else f"_{i+1}"
                if plain:
                    out[f"EMAIL{suffix}"] = val
                if sha:
                    out[f"EMAIL_SHA256{suffix}"] = sha256_hex(val)

        # Phone: same pattern.
        if m.get("PHONE_SOURCES"):
            collected = []
            for col in m["PHONE_SOURCES"]:
                raw = src.get(col, "") or ""
                for p in raw.split(sep):
                    n = norm_phone(p, prefix)
                    if n:
                        collected.append(n)
            seen = set()
            uniq = []
            for p in collected:
                if p not in seen:
                    seen.add(p)
                    uniq.append(p)
            uniq = uniq[:max_p]
            for i, val in enumerate(uniq):
                suffix = "" if i == 0 else f"_{i+1}"
                if plain:
                    out[f"PHONE{suffix}"] = val
                if sha:
                    out[f"PHONE_SHA256{suffix}"] = sha256_hex(val)

        if m.get("ZIP"):
            out["ZIP"] = norm_zip(src.get(m["ZIP"], ""))
        if m.get("CT"):
            out["CT"] = norm_city(src.get(m["CT"], ""))
        if m.get("ST"):
            out["ST"] = norm_state(src.get(m["ST"], ""))
        if m.get("COUNTRY"):
            out["COUNTRY"] = norm_country(src.get(m["COUNTRY"], ""))
        if m.get("DOB"):
            y, mo, d = parse_dob(src.get(m["DOB"], ""))
            out["DOBY"] = y
            out["DOBM"] = mo
            out["DOBD"] = d
        if m.get("GEN"):
            out["GEN"] = norm_gender(src.get(m["GEN"], ""))

        return out

    def _write_csv(self, cols: list) -> None:
        emit_every = max(1, min(5000, self.total_rows // 100 or 1))
        processed = 0
        with open(self.output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=cols, quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            for src in iter_input_rows(self.input_path):
                writer.writerow(self._build_row(src, cols))
                processed += 1
                if processed % emit_every == 0:
                    self.progress.emit(processed, max(self.total_rows, processed))
                    self.status.emit(f"Formatted {processed:,} rows...")
        self.progress.emit(processed, max(self.total_rows, processed))
        self.status.emit(
            f"Done. Wrote {processed:,} rows to {self.output_path}"
        )

    def _write_xlsx(self, cols: list) -> None:
        workbook = xlsxwriter.Workbook(
            self.output_path,
            {
                "constant_memory": True,
                "strings_to_formulas": False,
                "strings_to_urls": False,
                "strings_to_numbers": False,
            },
        )
        header_fmt = workbook.add_format({"bold": True})

        sheet_idx = 1
        sheet = workbook.add_worksheet(f"Audience_{sheet_idx}")
        for c, name in enumerate(cols):
            sheet.write_string(0, c, name, header_fmt)

        max_data_rows = EXCEL_MAX_ROWS - 1
        row_in_sheet = 1
        processed = 0
        emit_every = max(1, min(5000, self.total_rows // 100 or 1))

        for src in iter_input_rows(self.input_path):
            if row_in_sheet > max_data_rows:
                sheet_idx += 1
                sheet = workbook.add_worksheet(f"Audience_{sheet_idx}")
                for c, name in enumerate(cols):
                    sheet.write_string(0, c, name, header_fmt)
                row_in_sheet = 1
                self.warning.emit(
                    f"Row limit reached, continuing in Audience_{sheet_idx}."
                )
            row = self._build_row(src, cols)
            for c, name in enumerate(cols):
                v = row.get(name, "")
                if v:
                    sheet.write_string(row_in_sheet, c, v)
            row_in_sheet += 1
            processed += 1
            if processed % emit_every == 0:
                self.progress.emit(processed, max(self.total_rows, processed))
                self.status.emit(f"Formatted {processed:,} rows...")

        workbook.close()
        self.progress.emit(processed, max(self.total_rows, processed))
        self.status.emit(
            f"Done. Wrote {processed:,} rows to {self.output_path}"
        )


# =====================================================================
# Widgets
# =====================================================================


class DropZone(QFrame):
    file_dropped = pyqtSignal(str)
    clicked = pyqtSignal()

    def __init__(
        self,
        accept_extensions=(".csv",),
        title_text="Drop a .csv file here",
        sub_text="or use the Choose File button below",
    ):
        super().__init__()
        self.setObjectName("DropZone")
        self.setAcceptDrops(True)
        self.setMinimumHeight(160)
        self.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred
        )
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._accept = tuple(e.lower() for e in accept_extensions)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title = QLabel(title_text)
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        f = title.font()
        f.setPointSize(16)
        f.setBold(True)
        title.setFont(f)
        sub = QLabel(sub_text)
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sub.setStyleSheet("color: #9d9d9d;")
        layout.addWidget(title)
        layout.addWidget(sub)

        self._active = False
        self._restyle()

    def _accepts(self, path: str) -> bool:
        return path.lower().endswith(self._accept)

    def _restyle(self) -> None:
        border = "#007acc" if self._active else "#3e3e42"
        bg = "#2d2d30" if self._active else "#252526"
        self.setStyleSheet(
            f"QFrame#DropZone {{ background-color: {bg}; "
            f"border: 2px dashed {border}; border-radius: 8px; }}"
        )

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        urls = event.mimeData().urls() if event.mimeData().hasUrls() else []
        if any(self._accepts(u.toLocalFile()) for u in urls):
            event.acceptProposedAction()
            self._active = True
            self._restyle()

    def dragLeaveEvent(self, event) -> None:
        self._active = False
        self._restyle()

    def dropEvent(self, event: QDropEvent) -> None:
        for url in event.mimeData().urls():
            path = url.toLocalFile()
            if self._accepts(path):
                self.file_dropped.emit(path)
                break
        self._active = False
        self._restyle()

    def mousePressEvent(self, event) -> None:
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()


# =====================================================================
# Main window
# =====================================================================


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CSV to XLSX Converter")
        self.resize(960, 820)

        # Converter tab state
        self.input_path: Optional[str] = None
        self.output_path: Optional[str] = None
        self.row_count: int = 0
        self.warnings_log: list = []
        self.count_worker: Optional[RowCountWorker] = None
        self.convert_worker: Optional[ConvertWorker] = None

        # Formatter tab state
        self.fmt_input_path: Optional[str] = None
        self.fmt_output_path: Optional[str] = None
        self.fmt_columns: list = []
        self.fmt_worker: Optional[AudienceFormatWorker] = None

        tabs = QTabWidget()
        tabs.addTab(self._build_converter_tab(), "CSV → XLSX Converter")
        tabs.addTab(
            self._build_formatter_tab(), "Facebook Audience Formatter"
        )
        self.setCentralWidget(tabs)

    # ---------------------------------------------------------------
    # Converter tab
    # ---------------------------------------------------------------

    def _build_converter_tab(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)

        # Drop zone
        self.drop_zone = DropZone()
        self.drop_zone.file_dropped.connect(self.set_input_file)
        self.drop_zone.clicked.connect(self.choose_input_file)
        layout.addWidget(self.drop_zone)

        # File row
        file_row = QHBoxLayout()
        self.choose_btn = QPushButton("Choose File…")
        self.choose_btn.clicked.connect(self.choose_input_file)
        self.file_info = QLabel("No file selected")
        self.file_info.setStyleSheet("color: #9d9d9d;")
        file_row.addWidget(self.choose_btn)
        file_row.addWidget(self.file_info, 1)
        layout.addLayout(file_row)

        # Preview
        layout.addWidget(QLabel("Preview (first 5 rows):"))
        self.preview = QTableWidget(0, 0)
        self.preview.setMaximumHeight(170)
        self.preview.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive
        )
        self.preview.verticalHeader().setVisible(False)
        layout.addWidget(self.preview)

        # Output row
        out_row = QHBoxLayout()
        self.choose_out_btn = QPushButton("Choose Output Location…")
        self.choose_out_btn.clicked.connect(self.choose_output_path)
        self.out_label = QLabel("Output: (defaults to same folder as input)")
        self.out_label.setStyleSheet("color: #9d9d9d;")
        out_row.addWidget(self.choose_out_btn)
        out_row.addWidget(self.out_label, 1)
        layout.addLayout(out_row)

        # Convert
        self.convert_btn = QPushButton("Convert")
        self.convert_btn.setEnabled(False)
        self.convert_btn.clicked.connect(self.start_convert)
        f = self.convert_btn.font()
        f.setBold(True)
        f.setPointSize(f.pointSize() + 1)
        self.convert_btn.setFont(f)
        layout.addWidget(self.convert_btn)

        # Progress + status
        self.progress = QProgressBar()
        self.progress.setValue(0)
        layout.addWidget(self.progress)
        self.status = QLabel("Idle.")
        self.status.setWordWrap(True)
        layout.addWidget(self.status)

        # Warnings (collapsible)
        self.warnings_group = QGroupBox("Warnings (0)")
        self.warnings_group.setCheckable(True)
        self.warnings_group.setChecked(False)
        wlay = QVBoxLayout(self.warnings_group)
        self.warnings_view = QTextEdit()
        self.warnings_view.setReadOnly(True)
        self.warnings_view.setMaximumHeight(140)
        wlay.addWidget(self.warnings_view)
        self.warnings_group.toggled.connect(self.warnings_view.setVisible)
        self.warnings_view.setVisible(False)
        layout.addWidget(self.warnings_group)

        # Open folder (hidden until conversion succeeds)
        self.open_folder_btn = QPushButton("Open Output Folder")
        self.open_folder_btn.clicked.connect(self.open_output_folder)
        self.open_folder_btn.setVisible(False)
        layout.addWidget(self.open_folder_btn)

        return page

    # ---- converter handlers ----

    def choose_input_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Choose CSV",
            str(Path.home()),
            "CSV files (*.csv);;All files (*)",
        )
        if path:
            self.set_input_file(path)

    def set_input_file(self, path: str) -> None:
        if not os.path.isfile(path):
            QMessageBox.warning(self, "Not a file", f"Can't read: {path}")
            return
        self.input_path = path
        size = os.path.getsize(path)
        self.file_info.setText(
            f"{Path(path).name} — {human_size(size)} — counting rows…"
        )
        self.set_default_output(path)
        self._refresh_convert_enabled()

        if self.count_worker is not None and self.count_worker.isRunning():
            self.count_worker.requestInterruption()
        self.count_worker = RowCountWorker(path)
        self.count_worker.done.connect(self._on_count_done)
        self.count_worker.failed.connect(self._on_count_failed)
        self.count_worker.start()

    def _on_count_done(self, count: int, cols: list, rows: list) -> None:
        if not self.input_path:
            return
        self.row_count = count
        size = os.path.getsize(self.input_path)
        self.file_info.setText(
            f"{Path(self.input_path).name} — {human_size(size)} — {count:,} rows"
        )
        self._populate_preview(cols, rows)
        self._refresh_convert_enabled()

    def _on_count_failed(self, msg: str) -> None:
        self.row_count = 0
        if self.input_path:
            size = os.path.getsize(self.input_path)
            self.file_info.setText(
                f"{Path(self.input_path).name} — {human_size(size)} — row count unavailable"
            )
        self._log_warning(f"Could not count rows: {msg}")

    def _populate_preview(self, cols: list, rows: list) -> None:
        self.preview.clear()
        self.preview.setColumnCount(len(cols))
        self.preview.setHorizontalHeaderLabels([str(c) for c in cols])
        self.preview.setRowCount(len(rows))
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                self.preview.setItem(
                    r, c, QTableWidgetItem("" if val is None else str(val))
                )

    def set_default_output(self, input_path: str) -> None:
        self.output_path = str(Path(input_path).with_suffix(".xlsx"))
        self.out_label.setText(f"Output: {self.output_path}")

    def choose_output_path(self) -> None:
        default = self.output_path or str(Path.home() / "output.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save XLSX as", default, "Excel workbook (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        if not self._output_writable(path):
            QMessageBox.warning(
                self,
                "Cannot write here",
                "That folder isn't writable. Pick another location.",
            )
            return
        self.output_path = path
        self.out_label.setText(f"Output: {self.output_path}")
        self._refresh_convert_enabled()

    @staticmethod
    def _output_writable(path: str) -> bool:
        folder = os.path.dirname(path) or "."
        return os.access(folder, os.W_OK)

    def _refresh_convert_enabled(self) -> None:
        self.convert_btn.setEnabled(
            bool(self.input_path and self.output_path)
        )

    def start_convert(self) -> None:
        if not self.input_path or not self.output_path:
            return
        if not self._output_writable(self.output_path):
            QMessageBox.warning(
                self,
                "Output not writable",
                "Pick a different output location — that folder isn't writable.",
            )
            return

        self.warnings_log.clear()
        self.warnings_view.clear()
        self.warnings_group.setTitle("Warnings (0)")
        self.open_folder_btn.setVisible(False)
        self.convert_btn.setEnabled(False)
        self.choose_btn.setEnabled(False)
        self.choose_out_btn.setEnabled(False)

        if self.row_count > 0:
            self.progress.setRange(0, self.row_count)
            self.progress.setValue(0)
        else:
            self.progress.setRange(0, 0)  # indeterminate marquee
        self.status.setText("Starting conversion…")

        self.convert_worker = ConvertWorker(
            self.input_path, self.output_path, self.row_count
        )
        self.convert_worker.progress.connect(self._on_progress)
        self.convert_worker.status.connect(self.status.setText)
        self.convert_worker.warning.connect(self._log_warning)
        self.convert_worker.failed.connect(self._on_failed)
        self.convert_worker.done.connect(self._on_done)
        self.convert_worker.start()

    def _on_progress(self, processed: int, total: int) -> None:
        if total > 0:
            if self.progress.maximum() != total:
                self.progress.setRange(0, total)
            self.progress.setValue(min(processed, total))

    def _log_warning(self, msg: str) -> None:
        self.warnings_log.append(msg)
        self.warnings_view.append(msg)
        self.warnings_group.setTitle(
            f"Warnings ({len(self.warnings_log)})"
        )

    def _on_failed(self, msg: str) -> None:
        self.convert_btn.setEnabled(True)
        self.choose_btn.setEnabled(True)
        self.choose_out_btn.setEnabled(True)
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.status.setText("Conversion failed.")
        QMessageBox.critical(self, "Conversion failed", msg)

    def _on_done(self, output_path: str) -> None:
        self.convert_btn.setEnabled(True)
        self.choose_btn.setEnabled(True)
        self.choose_out_btn.setEnabled(True)
        if self.progress.maximum() == 0:
            self.progress.setRange(0, 1)
            self.progress.setValue(1)
        else:
            self.progress.setValue(self.progress.maximum())
        self.status.setText(f"Done. Saved to: {output_path}")
        self.open_folder_btn.setVisible(True)

    def open_output_folder(self) -> None:
        if not self.output_path:
            return
        subprocess.run(["open", os.path.dirname(self.output_path)])

    # ---------------------------------------------------------------
    # Facebook Audience Formatter tab
    # ---------------------------------------------------------------

    FIELD_LABELS = {
        "EXTERN_ID": "External ID (UUID)",
        "FN": "First name",
        "LN": "Last name",
        "ZIP": "Zip / Postal",
        "CT": "City",
        "ST": "State (2-char)",
        "COUNTRY": "Country (2-char)",
        "DOB": "Date of birth",
        "GEN": "Gender",
    }

    def _build_formatter_tab(self) -> QWidget:
        # The whole tab lives inside a scroll area so it always fits the
        # screen regardless of how many controls are visible.
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        intro = QLabel(
            "Format a contact list for Meta Custom Audience upload. "
            "Pick the source columns for emails and phones (multiple allowed) "
            "and which fields to pass through. Output emits plaintext and "
            "SHA256 columns side by side."
        )
        intro.setWordWrap(True)
        intro.setStyleSheet("color: #9d9d9d;")
        layout.addWidget(intro)

        # Compact drop zone (shorter than the converter's so it doesn't
        # eat half the screen).
        self.fmt_drop_zone = DropZone(
            accept_extensions=(".csv", ".xlsx", ".xlsm"),
            title_text="Drop a .csv or .xlsx file here",
            sub_text="or use the Choose File button below",
        )
        self.fmt_drop_zone.setMinimumHeight(100)
        self.fmt_drop_zone.setMaximumHeight(120)
        self.fmt_drop_zone.file_dropped.connect(self.fmt_set_input)
        self.fmt_drop_zone.clicked.connect(self.fmt_choose_input)
        layout.addWidget(self.fmt_drop_zone)

        # File row
        row = QHBoxLayout()
        self.fmt_choose_btn = QPushButton("Choose CSV/XLSX…")
        self.fmt_choose_btn.clicked.connect(self.fmt_choose_input)
        self.fmt_file_info = QLabel("No file selected")
        self.fmt_file_info.setStyleSheet("color: #9d9d9d;")
        row.addWidget(self.fmt_choose_btn)
        row.addWidget(self.fmt_file_info, 1)
        layout.addLayout(row)

        # ----- Email + Phone source columns (multi-select) -----
        sources_group = QGroupBox(
            "Email & Phone source columns (auto-detected; uncheck to exclude)"
        )
        sources_layout = QGridLayout(sources_group)
        sources_layout.setHorizontalSpacing(12)

        sources_layout.addWidget(QLabel("Email source columns:"), 0, 0)
        sources_layout.addWidget(QLabel("Phone source columns:"), 0, 1)

        self.fmt_email_list = QListWidget()
        self.fmt_email_list.setMaximumHeight(140)
        self.fmt_email_list.itemChanged.connect(self._fmt_update_source_counts)
        sources_layout.addWidget(self.fmt_email_list, 1, 0)

        self.fmt_phone_list = QListWidget()
        self.fmt_phone_list.setMaximumHeight(140)
        self.fmt_phone_list.itemChanged.connect(self._fmt_update_source_counts)
        sources_layout.addWidget(self.fmt_phone_list, 1, 1)

        self.fmt_email_count_label = QLabel("0 selected")
        self.fmt_email_count_label.setStyleSheet("color: #9d9d9d;")
        self.fmt_phone_count_label = QLabel("0 selected")
        self.fmt_phone_count_label.setStyleSheet("color: #9d9d9d;")
        sources_layout.addWidget(self.fmt_email_count_label, 2, 0)
        sources_layout.addWidget(self.fmt_phone_count_label, 2, 1)

        layout.addWidget(sources_group)

        # ----- Other fields (single-source dropdowns, two-column grid) -----
        fields_group = QGroupBox(
            "Other fields (set to '(none)' to skip in output)"
        )
        fields_grid = QGridLayout(fields_group)
        fields_grid.setHorizontalSpacing(12)
        fields_grid.setVerticalSpacing(6)
        self.fmt_combos: dict = {}
        # Lay out in two columns to keep vertical footprint small.
        for i, field in enumerate(FB_SINGLE_FIELDS):
            col = i // 5  # first 5 in column 0, rest in column 1
            row_idx = i % 5
            label = QLabel(self.FIELD_LABELS[field] + ":")
            cb = QComboBox()
            cb.setEnabled(False)
            cb.setMinimumWidth(160)
            self.fmt_combos[field] = cb
            fields_grid.addWidget(label, row_idx, col * 2)
            fields_grid.addWidget(cb, row_idx, col * 2 + 1)
        fields_grid.setColumnStretch(1, 1)
        fields_grid.setColumnStretch(3, 1)
        layout.addWidget(fields_group)

        # ----- Options (compact 2-column grid) -----
        opts_group = QGroupBox("Options")
        opts_grid = QGridLayout(opts_group)
        opts_grid.setHorizontalSpacing(12)
        opts_grid.setVerticalSpacing(6)

        self.fmt_separator = QComboBox()
        self.fmt_separator.addItems(
            [", (comma)", "; (semicolon)", "| (pipe)", "tab"]
        )
        self.fmt_max_emails = QSpinBox()
        self.fmt_max_emails.setRange(1, 20)
        self.fmt_max_emails.setValue(5)
        self.fmt_max_phones = QSpinBox()
        self.fmt_max_phones.setRange(1, 20)
        self.fmt_max_phones.setValue(4)
        self.fmt_country_prefix = QComboBox()
        self.fmt_country_prefix.setEditable(True)
        self.fmt_country_prefix.addItems(
            ["1 (US/CA)", "44 (UK)", "61 (AU)", "(none)"]
        )
        self.fmt_output_format = QComboBox()
        self.fmt_output_format.addItems(
            ["CSV (recommended for Meta upload)", "XLSX"]
        )
        self.fmt_output_format.currentIndexChanged.connect(
            self._fmt_on_format_change
        )

        opts_grid.addWidget(QLabel("Separator:"), 0, 0)
        opts_grid.addWidget(self.fmt_separator, 0, 1)
        opts_grid.addWidget(QLabel("Country code:"), 0, 2)
        opts_grid.addWidget(self.fmt_country_prefix, 0, 3)

        opts_grid.addWidget(QLabel("Max email cols:"), 1, 0)
        opts_grid.addWidget(self.fmt_max_emails, 1, 1)
        opts_grid.addWidget(QLabel("Max phone cols:"), 1, 2)
        opts_grid.addWidget(self.fmt_max_phones, 1, 3)

        opts_grid.addWidget(QLabel("Output format:"), 2, 0)
        opts_grid.addWidget(self.fmt_output_format, 2, 1, 1, 3)

        self.fmt_include_plaintext = QCheckBox("Include plaintext columns")
        self.fmt_include_plaintext.setChecked(True)
        self.fmt_include_sha256 = QCheckBox("Include SHA256 columns")
        self.fmt_include_sha256.setChecked(True)
        opts_grid.addWidget(self.fmt_include_plaintext, 3, 0, 1, 2)
        opts_grid.addWidget(self.fmt_include_sha256, 3, 2, 1, 2)

        opts_grid.setColumnStretch(1, 1)
        opts_grid.setColumnStretch(3, 1)
        layout.addWidget(opts_group)

        # ----- Output + action -----
        out_row = QHBoxLayout()
        self.fmt_choose_out_btn = QPushButton("Choose Output Location…")
        self.fmt_choose_out_btn.clicked.connect(self.fmt_choose_output)
        self.fmt_out_label = QLabel("Output: (default appears once a file is loaded)")
        self.fmt_out_label.setStyleSheet("color: #9d9d9d;")
        out_row.addWidget(self.fmt_choose_out_btn)
        out_row.addWidget(self.fmt_out_label, 1)
        layout.addLayout(out_row)

        self.fmt_format_btn = QPushButton("Format && Save")
        self.fmt_format_btn.setEnabled(False)
        self.fmt_format_btn.clicked.connect(self.fmt_start)
        f = self.fmt_format_btn.font()
        f.setBold(True)
        f.setPointSize(f.pointSize() + 1)
        self.fmt_format_btn.setFont(f)
        layout.addWidget(self.fmt_format_btn)

        self.fmt_progress = QProgressBar()
        self.fmt_progress.setValue(0)
        layout.addWidget(self.fmt_progress)
        self.fmt_status = QLabel("Idle.")
        self.fmt_status.setWordWrap(True)
        layout.addWidget(self.fmt_status)

        self.fmt_open_folder_btn = QPushButton("Open Output Folder")
        self.fmt_open_folder_btn.clicked.connect(self.fmt_open_output_folder)
        self.fmt_open_folder_btn.setVisible(False)
        layout.addWidget(self.fmt_open_folder_btn)

        layout.addStretch(1)
        scroll.setWidget(page)
        return scroll

    # ---- formatter handlers ----

    def fmt_choose_input(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Choose CSV or XLSX",
            str(Path.home()),
            "Data files (*.csv *.xlsx);;CSV (*.csv);;Excel (*.xlsx);;All files (*)",
        )
        if path:
            self.fmt_set_input(path)

    def fmt_set_input(self, path: str) -> None:
        if not os.path.isfile(path):
            QMessageBox.warning(self, "Not a file", f"Can't read: {path}")
            return
        ext = Path(path).suffix.lower()
        if ext not in (".csv", ".xlsx", ".xlsm"):
            QMessageBox.warning(
                self, "Unsupported", "Pick a .csv or .xlsx file."
            )
            return

        self.fmt_input_path = path
        size = os.path.getsize(path)
        self.fmt_file_info.setText(
            f"{Path(path).name} — {human_size(size)} — reading columns…"
        )
        try:
            cols = read_input_columns(path)
        except Exception as e:
            QMessageBox.critical(self, "Couldn't read file", str(e))
            self.fmt_file_info.setText("Couldn't read file.")
            return
        self.fmt_columns = cols
        self.fmt_file_info.setText(
            f"{Path(path).name} — {human_size(size)} — {len(cols)} columns detected"
        )
        self._fmt_populate_mapping(cols)
        self._fmt_set_default_output(path)
        self._fmt_refresh_enabled()

    def _fmt_populate_mapping(self, cols: list) -> None:
        # Single-value field dropdowns.
        for field, combo in self.fmt_combos.items():
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("(none)", userData="")
            for c in cols:
                combo.addItem(c, userData=c)
            combo.setEnabled(True)
            auto = auto_map_columns(field, cols)
            if auto:
                idx = combo.findData(auto)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
            combo.blockSignals(False)

        # Email & phone source checkable lists. Auto-check anything whose
        # name matches the pattern lists; the user can uncheck what they
        # don't want (e.g. unverified email columns).
        self._fmt_fill_source_list(
            self.fmt_email_list, cols, EMAIL_COLUMN_PATTERNS
        )
        self._fmt_fill_source_list(
            self.fmt_phone_list, cols, PHONE_COLUMN_PATTERNS
        )
        self._fmt_update_source_counts()

    @staticmethod
    def _fmt_fill_source_list(
        list_widget: QListWidget, cols: list, patterns: list
    ) -> None:
        list_widget.blockSignals(True)
        list_widget.clear()
        for c in cols:
            item = QListWidgetItem(c)
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            lower = c.lower()
            is_match = any(p in lower for p in patterns)
            item.setCheckState(
                Qt.CheckState.Checked if is_match else Qt.CheckState.Unchecked
            )
            list_widget.addItem(item)
        list_widget.blockSignals(False)

    def _fmt_update_source_counts(self, *_args) -> None:
        ec = self._fmt_checked_items(self.fmt_email_list)
        pc = self._fmt_checked_items(self.fmt_phone_list)
        self.fmt_email_count_label.setText(f"{len(ec)} selected")
        self.fmt_phone_count_label.setText(f"{len(pc)} selected")

    @staticmethod
    def _fmt_checked_items(list_widget: QListWidget) -> list:
        out = []
        for i in range(list_widget.count()):
            it = list_widget.item(i)
            if it.checkState() == Qt.CheckState.Checked:
                out.append(it.text())
        return out

    def _fmt_set_default_output(self, input_path: str) -> None:
        out_fmt = self._fmt_current_output_format()
        p = Path(input_path)
        default = p.with_name(p.stem + "_facebook." + out_fmt)
        self.fmt_output_path = str(default)
        self.fmt_out_label.setText(f"Output: {self.fmt_output_path}")

    def _fmt_current_output_format(self) -> str:
        return (
            "csv" if "CSV" in self.fmt_output_format.currentText() else "xlsx"
        )

    def _fmt_on_format_change(self) -> None:
        # Keep the default output path's extension in sync with the chosen format.
        if not self.fmt_output_path:
            return
        out_fmt = self._fmt_current_output_format()
        p = Path(self.fmt_output_path)
        new_path = str(p.with_suffix("." + out_fmt))
        self.fmt_output_path = new_path
        self.fmt_out_label.setText(f"Output: {self.fmt_output_path}")

    def fmt_choose_output(self) -> None:
        out_fmt = self._fmt_current_output_format()
        default = self.fmt_output_path or str(
            Path.home() / f"audience.{out_fmt}"
        )
        filter_str = (
            "CSV (*.csv)" if out_fmt == "csv" else "Excel workbook (*.xlsx)"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Save as", default, filter_str
        )
        if not path:
            return
        ext = "." + out_fmt
        if not path.lower().endswith(ext):
            path += ext
        if not self._output_writable(path):
            QMessageBox.warning(
                self,
                "Cannot write here",
                "That folder isn't writable. Pick another location.",
            )
            return
        self.fmt_output_path = path
        self.fmt_out_label.setText(f"Output: {self.fmt_output_path}")
        self._fmt_refresh_enabled()

    def _fmt_refresh_enabled(self) -> None:
        self.fmt_format_btn.setEnabled(
            bool(self.fmt_input_path and self.fmt_output_path)
        )

    def fmt_start(self) -> None:
        if not self.fmt_input_path or not self.fmt_output_path:
            return

        mapping: dict = {}
        for field, combo in self.fmt_combos.items():
            val = combo.currentData()
            if val:
                mapping[field] = val
        email_sources = self._fmt_checked_items(self.fmt_email_list)
        phone_sources = self._fmt_checked_items(self.fmt_phone_list)
        if email_sources:
            mapping["EMAIL_SOURCES"] = email_sources
        if phone_sources:
            mapping["PHONE_SOURCES"] = phone_sources

        sep_text = self.fmt_separator.currentText()
        separator = "\t" if "tab" in sep_text else sep_text.split(" ")[0]

        prefix_text = self.fmt_country_prefix.currentText()
        if "(none)" in prefix_text:
            country_prefix = ""
        else:
            # Take the first whitespace-separated token; works for both
            # the canned items ("1 (US/CA)") and free-typed values.
            country_prefix = re.sub(r"\D", "", prefix_text.split(" ")[0])

        options = {
            "max_emails": self.fmt_max_emails.value(),
            "max_phones": self.fmt_max_phones.value(),
            "include_plaintext": self.fmt_include_plaintext.isChecked(),
            "include_sha256": self.fmt_include_sha256.isChecked(),
            "separator": separator,
            "country_prefix": country_prefix,
            "output_format": self._fmt_current_output_format(),
        }

        if not options["include_plaintext"] and not options["include_sha256"]:
            QMessageBox.warning(
                self,
                "Nothing to output",
                "Enable at least one of plaintext or SHA256 columns.",
            )
            return

        if not mapping:
            QMessageBox.warning(
                self,
                "Nothing to output",
                "Map at least one field or check at least one email/phone "
                "source column.",
            )
            return

        # Sync output extension to format.
        expected_ext = "." + options["output_format"]
        if not self.fmt_output_path.lower().endswith(expected_ext):
            self.fmt_output_path = str(
                Path(self.fmt_output_path).with_suffix(expected_ext)
            )
            self.fmt_out_label.setText(f"Output: {self.fmt_output_path}")

        try:
            total = count_input_rows(self.fmt_input_path)
        except Exception:
            total = 0

        self.fmt_open_folder_btn.setVisible(False)
        self.fmt_format_btn.setEnabled(False)
        self.fmt_choose_btn.setEnabled(False)
        self.fmt_choose_out_btn.setEnabled(False)

        if total > 0:
            self.fmt_progress.setRange(0, total)
            self.fmt_progress.setValue(0)
        else:
            self.fmt_progress.setRange(0, 0)
        self.fmt_status.setText("Starting…")

        self.fmt_worker = AudienceFormatWorker(
            self.fmt_input_path,
            self.fmt_output_path,
            mapping,
            options,
            total,
        )
        self.fmt_worker.progress.connect(self._fmt_on_progress)
        self.fmt_worker.status.connect(self.fmt_status.setText)
        self.fmt_worker.failed.connect(self._fmt_on_failed)
        self.fmt_worker.done.connect(self._fmt_on_done)
        self.fmt_worker.start()

    def _fmt_on_progress(self, processed: int, total: int) -> None:
        if total > 0:
            if self.fmt_progress.maximum() != total:
                self.fmt_progress.setRange(0, total)
            self.fmt_progress.setValue(min(processed, total))

    def _fmt_on_failed(self, msg: str) -> None:
        self.fmt_format_btn.setEnabled(True)
        self.fmt_choose_btn.setEnabled(True)
        self.fmt_choose_out_btn.setEnabled(True)
        self.fmt_progress.setRange(0, 100)
        self.fmt_progress.setValue(0)
        self.fmt_status.setText("Formatting failed.")
        QMessageBox.critical(self, "Formatting failed", msg)

    def _fmt_on_done(self, output_path: str) -> None:
        self.fmt_format_btn.setEnabled(True)
        self.fmt_choose_btn.setEnabled(True)
        self.fmt_choose_out_btn.setEnabled(True)
        if self.fmt_progress.maximum() == 0:
            self.fmt_progress.setRange(0, 1)
            self.fmt_progress.setValue(1)
        else:
            self.fmt_progress.setValue(self.fmt_progress.maximum())
        self.fmt_status.setText(f"Done. Saved to: {output_path}")
        self.fmt_open_folder_btn.setVisible(True)

    def fmt_open_output_folder(self) -> None:
        if not self.fmt_output_path:
            return
        subprocess.run(["open", os.path.dirname(self.fmt_output_path)])


DARK_STYLESHEET = """
QMainWindow, QWidget { background-color: #1e1e1e; color: #e0e0e0; }
QLabel { color: #e0e0e0; }
QPushButton {
    background-color: #2d2d30; color: #e0e0e0;
    border: 1px solid #3e3e42; padding: 8px 14px; border-radius: 6px;
}
QPushButton:hover:!disabled { background-color: #3e3e42; border-color: #007acc; }
QPushButton:disabled { color: #6d6d6d; }
QProgressBar {
    border: 1px solid #3e3e42; border-radius: 4px; background-color: #2d2d30;
    text-align: center; color: #e0e0e0; height: 20px;
}
QProgressBar::chunk { background-color: #007acc; border-radius: 3px; }
QTextEdit {
    background-color: #1e1e1e; color: #d4d4d4; border: 1px solid #3e3e42;
    font-family: 'Menlo', monospace; font-size: 11px;
}
QTableWidget {
    background-color: #252526; color: #e0e0e0;
    border: 1px solid #3e3e42; gridline-color: #3e3e42;
}
QHeaderView::section {
    background-color: #2d2d30; color: #e0e0e0;
    border: 1px solid #3e3e42; padding: 4px;
}
QGroupBox {
    border: 1px solid #3e3e42; border-radius: 4px;
    margin-top: 10px; padding-top: 12px; color: #e0e0e0;
}
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }
QTabWidget::pane { border: 1px solid #3e3e42; top: -1px; }
QTabBar::tab {
    background: #2d2d30; color: #e0e0e0;
    padding: 8px 14px; border: 1px solid #3e3e42; border-bottom: none;
    border-top-left-radius: 6px; border-top-right-radius: 6px;
    margin-right: 2px;
}
QTabBar::tab:selected { background: #1e1e1e; border-color: #007acc; }
QTabBar::tab:!selected:hover { background: #3e3e42; }
QComboBox, QSpinBox {
    background-color: #2d2d30; color: #e0e0e0;
    border: 1px solid #3e3e42; padding: 4px 8px; border-radius: 4px;
    min-height: 22px;
}
QComboBox:hover, QSpinBox:hover { border-color: #007acc; }
QComboBox QAbstractItemView {
    background-color: #2d2d30; color: #e0e0e0;
    selection-background-color: #007acc;
}
QCheckBox { color: #e0e0e0; spacing: 6px; }
"""


def main() -> None:
    app = QApplication(sys.argv)
    app.setApplicationName("CSV to XLSX Converter")
    app.setStyleSheet(DARK_STYLESHEET)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
